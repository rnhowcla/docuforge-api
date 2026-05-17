import io
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd


def clean_excel(file_bytes: bytes, remove_duplicates: bool = True, trim_spaces: bool = True) -> bytes:
    df = pd.read_excel(io.BytesIO(file_bytes))
    if trim_spaces:
        for col in df.select_dtypes(include=["object", "string"]).columns:
            df[col] = df[col].str.strip()
    if remove_duplicates:
        df = df.drop_duplicates()
    output = io.BytesIO()
    df.to_excel(output, index=False)
    output.seek(0)
    return output.read()


def excel_to_csv(file_bytes: bytes, sheet_name: str | None = None) -> str:
    dfs = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name)
    if isinstance(dfs, dict):
        result = []
        for name, df in dfs.items():
            result.append(f"# Sheet: {name}")
            result.append(df.to_csv(index=False))
            result.append("")
        return "\n".join(result)
    return dfs.to_csv(index=False)


def format_excel(file_bytes: bytes) -> bytes:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)
        header_fill = openpyxl.styles.PatternFill("solid", fgColor="E0E0E0")
        header_font = openpyxl.styles.Font(bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
